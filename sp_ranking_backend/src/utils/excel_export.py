from typing import Iterable, Dict, Any, List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

# PUBLIC_INTERFACE
def build_excel_from_items(items: Iterable[Dict[str, Any]], sheet_name: str = "Rankings") -> bytes:
    """Create an Excel workbook from a list of dict items and return bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    items_list: List[Dict[str, Any]] = list(items or [])
    if not items_list:
        # Write a placeholder header
        ws.append(["No data"])
    else:
        headers = []
        # Collect headers from all items (stable order)
        seen = set()
        for row in items_list:
            for k in row.keys():
                if k not in seen:
                    seen.add(k)
                    headers.append(k)
        ws.append(headers)
        for row in items_list:
            ws.append([row.get(h) for h in headers])

        # Auto width
        for ci, h in enumerate(headers, start=1):
            col = get_column_letter(ci)
            max_len = max(len(str(h)), *(len(str(r.get(h))) for r in items_list if r.get(h) is not None))
            ws.column_dimensions[col].width = min(60, max_len + 2)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
